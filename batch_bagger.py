# -*- coding: utf-8 -*-

import bagit
import sys
import uuid
import csv
import os
import re
import argparse
import shutil

parser = argparse.ArgumentParser(description='Bag a list of folders in a directory, substituting values from a CSV into a bag-info template')
parser.add_argument('-d', '--directory', help="directory containing folders to be bagged", action="store", dest="d", default=os.getcwd())
parser.add_argument('-b', '--baginfo', required=False, help="bag-info.txt file to be distributed into each bag", action="store", dest="b")
parser.add_argument('-s', '--spreadsheet', required=False, help="CSV or XLSX file containing the names of folders to be bagged, and any values to be substituted in the bag-info file", action="store", dest="s")
parser.add_argument('-v', '--verbose', required=False, help="print the contents of each bag-info file as it is written", action="store_true", dest="v")
parser.add_argument('-u', '--unpack', required=False, help="unpack existing bags' data directories", action="store_true", dest="u")

args = parser.parse_args()

bagsDir = args.d

# Allow users to add a -v switch, which will print each bag-info file as it is written
if args.v:
    verbose = True
else:
    verbose = False

    
# Define the text strings that will be recognized as bag-info field labels
fieldsList = ['Source-Organization', 'Organization-Address', 'Contact-Name', 'Contact-Phone', 'Contact-Email', 'External-Description', 
              'External-Identifier', 'Internal-Sender-Description', 'Internal-Sender-Identifier', 'Rights-Statement', 'Bag-Group-Identifier', 'Bag-Size']

def bagCreator(bagPath):
    
    # Store the contents of each row, and create a dictionary of column
    # headers and values that will replace placeholder text in the bag-info
    # template file
    replacementDict = {}
    for i in range(len(replaceFields)):
        replacementDict[replaceFields[i]] = rowList[i]

    # Iterate over the template file, assembling a bag-info dictionary for each entry
    baginfoDict = {}
    # Generate a UUID to be used as an External-Identifier later
    bagUUID = str(uuid.uuid4())
    with open(bagInfo, 'r', encoding='utf-8') as templateFile:
        for line in templateFile:
            line = line.lstrip()
            
            # Skip blank lines, to avoid duplicating the final field in the 
            # bag-info template
            if line == '':
                next
                
            # If the line contains a colon, check, to see if the line begins
            # with one of the labels found in the fieldsList
            if line.split(':')[0] in fieldsList:
                    # If the label is already a key in the bag-info dictionary, 
                    # append the new value to the existing value, to combine 
                    # values of repeated fields
                    if line.split(':')[0] in baginfoDict:
                        label = line.split(':')[0]
                        value = ':'.join(line.split(':')[1:])
                        value = value.lstrip()
                        baginfoDict[label] = baginfoDict[label] + ' | ' + value
                        
                    # If the label is not already in the dictionary, create a new
                    # dictionary entry with the label and value
                    else:
                        label = line.split(':')[0]
                        value = ':'.join(line.split(':')[1:])
                        value = value.lstrip()
                        baginfoDict[label] = value
            
            # If the text before the colon is not found in the fieldsList,
            # it is a continuation of the previous line's text, and just
            # happens to contain a colon. In this case, add the line to 
            # the previous dictionary entry                
            elif not line.split(':')[0] in fieldsList:
                baginfoDict[label] = baginfoDict[label] + line
    
    # Add the UUID External-Identifier to the bag-info dictionary. If other
    # External-Identifier values are present, insert the UUID first.
    if 'External-Identifier' in baginfoDict:
        baginfoDict['External-Identifier'] = bagUUID + ' | ' + baginfoDict['External-Identifier']
    else:
        baginfoDict['External-Identifier'] = bagUUID
        
    # Iterate over the bag-info dictionary, replacing keywords with corresponding
    # values from the spreadsheet. Also remove trailing white space in each value,
    # replace line break characters with spaces, and replace double spaces with
    # single spaces.
    for key, value in baginfoDict.items():
        value = value.rstrip()
        value = value.replace('\n', ' ')
        value = value.replace('  ', ' ')
        
        # Look for text in the bag-info dictionary wrapped in double brackets,
        # matching the column headers in the spreadsheet. Replace that text with
        # the corresponding text in the spreadsheet. Update the bag-info dictionary
        # with the updated text.
        for label, replacementValue in replacementDict.items():
            value = re.sub('\[\[' + label + '\]\]', replacementValue, value, flags=re.IGNORECASE)
            baginfoDict[key] = value
    
    # Set the UUIDs.txt file path
    outPath = os.path.join(bagsDir, 'UUIDs.csv')
    
    # Name the bag being created
    print('Bagging: ' + rowList[0])
    
    # If the user has added a -v switch, print the bag-info dictionary
    if verbose:
        for key in baginfoDict:
            print(key + ': ' + baginfoDict[key])
            
    # Create the bag at the bag directory, using the first column of the spreadsheet                
    bagPath = os.path.join(bagsDir, rowList[0])
    
    # Calculate the bag size
    baginfoDict['Bag-Size'] = sizeCalculator(bagPath)

    bag = bagit.make_bag(bagPath, baginfoDict, checksum=['sha256']) # Specify a sha256 hash algorithm
    
    # After the bag has been written, write the UUIDs.txt file
    with open(outPath, 'a+', encoding='utf-8', newline='') as guids_file:
        uuidWriter = csv.writer(guids_file, delimiter=',', quotechar='"')
        outRow = [rowList[0], bagUUID]
        uuidWriter.writerow(outRow)

def bagUnpacker(bag):
    os.remove(os.path.join(bag, 'bag-info.txt'))
    os.remove(os.path.join(bag, 'bagit.txt'))
    os.remove(os.path.join(bag, 'manifest-sha256.txt'))
    os.remove(os.path.join(bag, 'tagmanifest-sha256.txt'))
    dataPath = os.path.join(bag,'data')
    for item in os.listdir(dataPath):
        if not item == 'data':
            shutil.move(os.path.join(dataPath, item), bag)
        elif item == 'data':
            # Move any items named "data" and give them a temporary name before 
            # removing "dataPath" and renaming the temporary item "data"
            origPath = os.path.join(dataPath, item)
            tempPath = os.path.join(bag, (item + '_temp'))
            os.rename(origPath, tempPath)
    os.rmdir(dataPath)
    try:
        os.rename(tempPath, dataPath)
    except:
        next
    print('Unpacked: ' + bag)

def sizeCalculator(bag):
    total = 0
    for root, dirs, files in os.walk(bag):
        for file in files:
            filepath = os.path.join(root, file)
            total += os.path.getsize(filepath)
    kbTotal = total / 1024
    mbTotal = total / 1048576
    gbTotal = total / 1073741824
    tbTotal = total / 1099511627776
    
    if tbTotal > 1:
        return(str(round(tbTotal, 1)) + ' TB')
    if gbTotal > 1:
        return(str(round(gbTotal, 1)) + ' GB')
    elif mbTotal > 1:
        return(str(round(mbTotal, 1)) + ' MB')
    elif kbTotal > 1:
        return(str(round(kbTotal, 1)) + ' KB')
    else:
        return(str(total) + ' bytes')

if args.u:
    try:
        print("WARNING: This will unpack all bags at the target directory, moving the contents of the data directory up one level and deleting the bag-info, bagit, manifest, and tagmanifest files. This will INVALIDATE bags. Do not run if you are not absolutely certain that this is what you want to do!") 
        input("Press Enter here to continue, or CTRL+C to abort...")
    except KeyboardInterrupt:
        sys.exit('\nScript aborted by user.')
    if args.s:
        spreadSheet = args.s
        if spreadSheet.endswith('.csv'):
            with open(spreadSheet, 'r', encoding='utf-8') as bag_spreadsheet:
                spreadsheet_reader = csv.reader(bag_spreadsheet, delimiter=',', quotechar='"')
                for entry in spreadsheet_reader: 
                    # Iterate over the spreadsheet, skipping the column headers
                    if spreadsheet_reader.line_num == 0:
                        next
                    bagPath = os.path.join(bagsDir, entry[0])
                    bagUnpacker(bagPath)
        elif spreadSheet.endswith('.xlsx'):
            
            from openpyxl import load_workbook
            
            wb = load_workbook(spreadSheet, read_only=True)
            sheetName = wb.sheetnames[0]
            ws = wb[sheetName]
            
            for row in ws[2: ws.max_row]:
                rowList = []
                for cell in row:
                    rowList.append(cell.value)
                bagPath = os.path.join(bagsDir, (rowList[0]))
                bagUnpacker(bagPath)
                
    else:
        for item in os.listdir(bagsDir):
            if os.path.isdir(item):
                # Test whether a given subfolder is actually a bag and should
                # be unpacked. When no spreadsheet is provided, this will 
                # prevent the bagUnpacker function from running on non-bag 
                # directories, which would throw an error
                dirContents = os.listdir(item)
                if 'bag-info.txt' in dirContents:
                    bagPath = os.path.join(bagsDir,item)
                    bagUnpacker(bagPath)
                else:
                    print('Directory ' + item + ' is not a bag, skipping...')
    
if not args.u:
    if not args.s:
        sys.exit('Please provide a CSV or XLSX bag-info spreadsheet and run again.')
    else:
        spreadSheet = args.s
    
    if not args.b:
        sys.exit('Please provide a template bag-info file and run again.')
    else:
        bagInfo = args.b
        
    if spreadSheet.endswith('.csv'):
        with open(spreadSheet, 'r', encoding='utf-8') as bag_spreadsheet:
            spreadsheet_reader = csv.reader(bag_spreadsheet, delimiter=',', quotechar='"')
            
            # Store the column headers in the first line of the spreadsheet
            replaceFields = spreadsheet_reader.__next__()
            
            for entry in spreadsheet_reader: 
                # Iterate over the spreadsheet, skipping the column headers
                if spreadsheet_reader.line_num == 0:
                    next
    
                rowList = entry
                bagCreator(rowList[0])
    
    elif spreadSheet.endswith('.xlsx'):
        
        from openpyxl import load_workbook
        
        # Open the workbook at read only
        wb = load_workbook(spreadSheet, read_only=True)
        sheetName = wb.sheetnames[0]
        ws = wb[sheetName]
        
        # Store the column headers in the first line of the spreadsheet
        firstrow = ws[1]
        replaceFields = []
        for cell in firstrow:
            replaceFields.append(cell.value)
        
        # Iterate over the remaining rows, storing the cell values in a list
        for row in ws[2: ws.max_row]:
            rowList = []
            for cell in row:
                rowList.append(cell.value)
            bagCreator(rowList[0])
